import pdfplumber
import pandas as pd
import re
import os
import logging

from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIGURAÇÕES
# =========================================================

PASTA = Path.home() / "Downloads" / "notas"

log_arquivo = PASTA / "processamento.log"

logging.basicConfig(
    filename=log_arquivo,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    encoding="utf-8",
)

# =========================================================
# FUNÇÕES AUXILIARES
# =========================================================


def buscar(regex, texto, grupo=1):

    resultado = re.search(regex, texto, re.IGNORECASE | re.DOTALL)

    if resultado:
        return resultado.group(grupo).strip()

    return ""


def formatar_data(data_texto):

    if not data_texto:
        return ""

    try:

        return datetime.strptime(data_texto, "%d.%m.%Y").strftime("%d/%m/%Y")

    except Exception:

        return data_texto.replace(".", "/")


def ler_pdf(caminho_pdf):

    paginas = []

    try:

        with pdfplumber.open(caminho_pdf) as pdf:

            for pagina in pdf.pages:

                try:

                    texto = pagina.extract_text()

                    if texto:
                        paginas.append(texto)

                except Exception as erro_pagina:

                    logging.error(
                        f"Erro na página do PDF " f"{caminho_pdf.name}: {erro_pagina}"
                    )

        return "\n".join(paginas)

    except Exception as erro_pdf:

        logging.error(f"Erro ao abrir PDF " f"{caminho_pdf.name}: {erro_pdf}")

        return ""


# =========================================================
# EXTRAÇÃO DOS DADOS
# =========================================================


def extrair_dados(texto):

    score = 0

    suspeitas = []
    campos_faltando = []

    # =====================================================
    # OC
    # =====================================================

    oc = buscar(r"OC:\s*([\d\s]+?)\s*BIP", texto)

    oc = re.sub(r"\s+", "", oc)

    if oc:

        score += 30

        if len(oc) < 6:
            suspeitas.append("OC muito curta")

    else:
        campos_faltando.append("OC")

    # =====================================================
    # NOTA
    # =====================================================

    nota = buscar(r"Nº[:\s]*0*([0-9]+)", texto)

    if nota:

        score += 20

    else:
        campos_faltando.append("Nº")

    # =====================================================
    # DATA
    # =====================================================

    data_original = buscar(r"DATA DA EMISSÃO.*?([0-9]{2}\.[0-9]{2}\.[0-9]{4})", texto)

    data_formatada = formatar_data(data_original)

    if data_formatada:

        score += 10

    else:
        campos_faltando.append("DATA")

    # =====================================================
    # QUANTIDADE
    # =====================================================

    quantidade = buscar(r"(?:QUANTIDADE|QTD|QTDE|QUANT\.?)\D*([0-9]+)", texto)

    if quantidade:

        try:

            qtd_int = int(quantidade)

            if qtd_int > 0:

                score += 20

            else:

                suspeitas.append("Quantidade inválida")

        except:

            suspeitas.append("Quantidade inválida")

    else:
        campos_faltando.append("QUANTIDADE")

    # =====================================================
    # BIP
    # =====================================================

    bip = ""

    padroes_bip = [
        r"BIP\s*N[º°o]?\s*:\s*0*([0-9]+)",
        r"BIP\s*N[º°o]?\s*0*([0-9]+)",
    ]

    for padrao in padroes_bip:

        resultado = re.search(padrao, texto, re.IGNORECASE)

        if resultado:

            bip = resultado.group(1)

            break

    if bip:

        score += 20

    else:

        campos_faltando.append("BIP Nº")

    # =====================================================
    # OBSERVAÇÕES
    # =====================================================

    observacoes = []

    if campos_faltando:

        observacoes.append("Faltando: " + ", ".join(campos_faltando))

    if suspeitas:

        observacoes.append("Suspeitas: " + ", ".join(suspeitas))

    observacao = " | ".join(observacoes)

    # =====================================================
    # STATUS
    # =====================================================

    if campos_faltando:

        status = "ERRO"

    elif suspeitas or score < 80:

        status = "SUSPEITO"

    else:

        status = "OK"

    return {
        "OC": oc,
        "Nº": nota,
        "DATA DA EMISSÃO": data_formatada,
        "QUANTIDADE": quantidade,
        "BIP Nº": bip,
        "STATUS": status,
        "CONFIANÇA": f"{score}%",
        "OBSERVAÇÃO": observacao,
    }


# =========================================================
# AJUSTAR EXCEL
# =========================================================


def ajustar_excel(caminho_excel):

    wb = load_workbook(caminho_excel)

    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    amarelo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    vermelho = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for ws in wb.worksheets:

        # Cabeçalho
        for cell in ws[1]:

            cell.font = Font(bold=True)

        # Congelar linha
        ws.freeze_panes = "A2"

        # Filtro
        ws.auto_filter.ref = ws.dimensions

        # Ajustar largura
        for coluna in ws.columns:

            maior = 0

            letra_coluna = get_column_letter(coluna[0].column)

            for cell in coluna:

                try:

                    tamanho = len(str(cell.value))

                    if tamanho > maior:
                        maior = tamanho

                except:
                    pass

            ws.column_dimensions[letra_coluna].width = maior + 5

        # Colorir status
        for linha in ws.iter_rows(min_row=2):

            for cell in linha:

                if cell.value == "OK":
                    cell.fill = verde

                elif cell.value == "SUSPEITO":
                    cell.fill = amarelo

                elif cell.value == "ERRO":
                    cell.fill = vermelho

    wb.save(caminho_excel)


# =========================================================
# PROCESSAMENTO PRINCIPAL
# =========================================================


def main():

    print("\n🔍 Iniciando processamento...\n")

    logging.info("Processamento iniciado")

    # =====================================================
    # VERIFICAR PASTA
    # =====================================================

    if not PASTA.exists():

        print(f"❌ Pasta não encontrada:\n{PASTA}")

        logging.error(f"Pasta não encontrada: {PASTA}")

        return

    arquivos_pdf = sorted(
        [arq for arq in os.listdir(PASTA) if arq.lower().endswith(".pdf")]
    )

    if not arquivos_pdf:

        print("⚠️ Nenhum PDF encontrado.")

        logging.warning("Nenhum PDF encontrado.")

        return

    # =====================================================
    # LISTAS
    # =====================================================

    notas_processadas = []
    suspeitos = []
    erros_encontrados = []

    total = len(arquivos_pdf)

    # =====================================================
    # LOOP DOS PDFs
    # =====================================================

    for indice, arquivo in enumerate(arquivos_pdf, start=1):

        print(f"📄 [{indice}/{total}] {arquivo}")

        caminho_pdf = PASTA / arquivo

        try:

            texto = ler_pdf(caminho_pdf)

            # PDF sem texto
            if not texto.strip():

                erros_encontrados.append(
                    {
                        "ARQUIVO": arquivo,
                        "STATUS": "ERRO",
                        "OBSERVAÇÃO": "PDF sem texto detectável",
                    }
                )

                erros_encontrados.append({})

                continue

            # Extrair dados
            dados = extrair_dados(texto)

            # =================================================
            # OK
            # =================================================

            if dados["STATUS"] == "OK":

                notas_processadas.append(
                    {
                        "OC": dados["OC"],
                        "Nº": dados["Nº"],
                        "DATA DA EMISSÃO": dados["DATA DA EMISSÃO"],
                        "QUANTIDADE": dados["QUANTIDADE"],
                        "BIP Nº": dados["BIP Nº"],
                        "CONFIANÇA": dados["CONFIANÇA"],
                        "STATUS": dados["STATUS"],
                    }
                )

                # linha vazia
                notas_processadas.append({})

            # =================================================
            # SUSPEITO
            # =================================================

            elif dados["STATUS"] == "SUSPEITO":

                suspeitos.append(
                    {
                        "ARQUIVO": arquivo,
                        "OC": dados["OC"],
                        "Nº": dados["Nº"],
                        "DATA DA EMISSÃO": dados["DATA DA EMISSÃO"],
                        "QUANTIDADE": dados["QUANTIDADE"],
                        "BIP Nº": dados["BIP Nº"],
                        "CONFIANÇA": dados["CONFIANÇA"],
                        "STATUS": dados["STATUS"],
                        "OBSERVAÇÃO": dados["OBSERVAÇÃO"],
                    }
                )

                suspeitos.append({})

            # =================================================
            # ERRO
            # =================================================

            else:

                erros_encontrados.append(
                    {
                        "ARQUIVO": arquivo,
                        "OC": dados["OC"],
                        "Nº": dados["Nº"],
                        "DATA DA EMISSÃO": dados["DATA DA EMISSÃO"],
                        "QUANTIDADE": dados["QUANTIDADE"],
                        "BIP Nº": dados["BIP Nº"],
                        "CONFIANÇA": dados["CONFIANÇA"],
                        "STATUS": dados["STATUS"],
                        "OBSERVAÇÃO": dados["OBSERVAÇÃO"],
                    }
                )

                erros_encontrados.append({})

            logging.info(f"PDF processado: {arquivo}")

        except Exception as erro:

            erros_encontrados.append(
                {"ARQUIVO": arquivo, "STATUS": "ERRO", "OBSERVAÇÃO": str(erro)}
            )

            erros_encontrados.append({})

            logging.error(f"Erro geral no arquivo " f"{arquivo}: {erro}")

    # =====================================================
    # DATAFRAMES
    # =====================================================

    df_ok = pd.DataFrame(notas_processadas)

    df_suspeitos = pd.DataFrame(suspeitos)

    df_erros = pd.DataFrame(erros_encontrados)

    # =====================================================
    # SALVAR EXCEL
    # =====================================================

    nome_excel = f"controle_notas_" f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    caminho_excel = PASTA / nome_excel

    try:

        with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:

            df_ok.to_excel(writer, sheet_name="NOTAS_PROCESSADAS", index=False)

            df_suspeitos.to_excel(writer, sheet_name="SUSPEITOS", index=False)

            df_erros.to_excel(writer, sheet_name="ERROS_ENCONTRADOS", index=False)

        ajustar_excel(caminho_excel)

        print("\n✅ PROCESSAMENTO FINALIZADO")

        print(f"\n📁 Excel salvo em:\n" f"{caminho_excel}")

        print(f"\n✅ Notas OK: " f"{len(df_ok)}")

        print(f"⚠️ Suspeitos: " f"{len(df_suspeitos)}")

        print(f"❌ Erros: " f"{len(df_erros)}")

        logging.info(f"Excel criado com sucesso: " f"{caminho_excel}")

    except Exception as erro_excel:

        print("\n❌ Erro ao salvar Excel.")

        logging.error(f"Erro ao salvar Excel: " f"{erro_excel}")


# =========================================================
# EXECUÇÃO
# =========================================================

if __name__ == "__main__":
    main()
